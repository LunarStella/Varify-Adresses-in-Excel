import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#참고
#네이버 지도에서 검색 시 존재 확인 




#파일 이름
#엑셀 속성에서 위치 확인 -> 뒤에 /엑셀 이름.xlsx 추가, 반드시 r'파일 위치' 형식으로 기입!
# ex) r'C:\Users\Desktop\data.xlsx'   (여기서 '\' 는 '/'와 다르다. 키보드 원화 표시 찾기)
current_file = '.xlsx'

#엑셀 시트 이름, 반드시 ' ' 안에 시트 이름 기입!
current_sheet = 'Sheet1'

#(열(column) : 세로줄,  행(row) : 가로줄 )
#검사 할 열 번호 설정
column = 

#검사 할 행의 시작 번호 설정 
start_row = 

#검사 할 행의 끝 번호 설정
end_row = 

#기입 후 f5 혹은 상단 실행 -> 디버깅 시작 버튼 누르면 프로그램 작동
#반드시 엑셀을 닫고 실행!! (켜고 실행 시 error 발생)




#하단 data 임의로 수정 금지 
def validate_address(address, client_id, client_secret):
    # 네이버지도 Geocoding API 요청 URL 생성
    url = 'https://naveropenapi.apigw.ntruss.com/map-geocode/v2/geocode?query={}'.format(address)
    
    # 헤더에 인증 정보 추가
    headers = {'X-NCP-APIGW-API-KEY-ID': client_id, 'X-NCP-APIGW-API-KEY': client_secret}
    
    # 요청 보내기
    response = requests.get(url, headers=headers)
    
    # 응답에서 유효한 주소인지 확인
    if response.ok:
        json_data = response.json()
        if json_data['meta']['count'] > 0:
            if json_data['addresses'][0]['roadAddress'] != '':
                return True
    return False


# 엑셀 파일 로드
workbook = load_workbook(filename=current_file)
# Sheet1 시트 선택
sheet = workbook[current_file]


df = pd.read_excel(current_file, sheet_name = current_sheet, usecols = [column-1], header = None, engine = 'openpyxl')

#네이버 api 아이디
client_id = ''
client_secret = ''

#(n, m) -> n, n+1, n+2, ..., m-1 (m은 포함 x)
for current_row in range(start_row-1, end_row):
    #df.iloc[행의 번호, 열의 번호] (행 : 가로줄 열 : 세로줄)
    is_valid = validate_address(df.iloc[current_row,0], client_id, client_secret)
    #셀 선택
    cell = sheet.cell(current_row+1,column)
    #배경색 변경
    #주소가 비정상이면 빨강, 주소가 정상이면 초록
    if is_valid:
        cell.fill = PatternFill(patternType='solid', fgColor='81c147')
    else:
        cell.fill = PatternFill(patternType='solid', fgColor='FF0000')

    print("현재 검사한 행: ", current_row+1)

 # True

workbook.save(filename=current_file)

print("finished")





